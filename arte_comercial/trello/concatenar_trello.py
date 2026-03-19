import os
import subprocess
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows



BASE_DIR = r"C:\Users\pietr\OneDrive\.vscode\arte_\DOWNLOADS"
PATH_HEAVY = os.path.join(BASE_DIR, "master_heavy.xlsx")
PATH_ULTRA = os.path.join(BASE_DIR, "master_heavy_ultra.xlsx")
SCRIPT_ARCHIVE = r"C:\Users\pietr\OneDrive\.vscode\arte_\scripts\trello\trello_achive.py"


def clean_number(value):
    if pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).replace("R$", "").replace(" ", "").strip()
    if not text:
        return 0.0

    try:
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        elif "," in text:
            text = text.replace(",", ".")
        return float(text)
    except ValueError:
        return 0.0


def ensure_valor_venda_total(df):
    if "QTDE" not in df.columns or "VALOR_VENDA" not in df.columns:
        return df

    df = df.copy()
    qtde = df["QTDE"].apply(clean_number)
    valor_venda = df["VALOR_VENDA"].apply(clean_number)
    df["VALOR_VENDA_TOTAL"] = (qtde * valor_venda).round(2)
    return df


def insert_col_after(columns, col_name, anchor_col):
    ordered = [col for col in columns if col != col_name]
    idx = ordered.index(anchor_col) + 1 if anchor_col in ordered else len(ordered)
    ordered.insert(idx, col_name)
    return ordered


def concatenar_planilhas():
    print("--- Iniciando sincronizacao: Heavy -> Ultra (preservando headers) ---")

    if not os.path.exists(PATH_HEAVY):
        print(f"Erro: arquivo {PATH_HEAVY} nao encontrado.")
        return False

    if not os.path.exists(PATH_ULTRA):
        print(f"Erro: arquivo {PATH_ULTRA} nao encontrado. O Ultra deve existir para preservar a ordem.")
        return False

    df_heavy = ensure_valor_venda_total(pd.read_excel(PATH_HEAVY, header=1))
    df_ultra = ensure_valor_venda_total(pd.read_excel(PATH_ULTRA))

    original_cols = df_ultra.columns.tolist()
    if "VALOR_VENDA" in original_cols or "VALOR_VENDA" in df_heavy.columns:
        original_cols = insert_col_after(original_cols, "VALOR_VENDA_TOTAL", "VALOR_VENDA")

    print(f"Heavy: {len(df_heavy)} linhas carregadas.")
    print(f"Ultra: {len(df_ultra)} linhas carregadas (ordem com {len(original_cols)} colunas preservada).")

    if "ARQUIVO" not in df_heavy.columns or "Nº" not in df_heavy.columns:
        print("Erro: colunas 'ARQUIVO' ou 'Nº' ausentes no Heavy.")
        return False

    df_heavy["temp_key"] = df_heavy["ARQUIVO"].astype(str) + "_" + df_heavy["Nº"].astype(str)
    df_ultra["temp_key"] = df_ultra["ARQUIVO"].astype(str) + "_" + df_ultra["Nº"].astype(str)

    cols_to_sync = [
        "marca_modelo_edital",
        "STATUS",
        "marca_sugerida",
        "modelo_sugerido",
        "VALOR",
        "VALOR_VENDA",
        "VALOR_VENDA_TOTAL",
        "MARGEM_LUCRO",
        "LUCRO_TOTAL",
        "JUSTIFICATIVA_TECNICA",
        "PARECER_JURIDICO_IMPUGNACAO",
    ]
    cols_to_sync = [col for col in cols_to_sync if col in df_heavy.columns]

    mask_exists = df_heavy["temp_key"].isin(df_ultra["temp_key"])
    to_update = df_heavy[mask_exists]
    to_add = df_heavy[~mask_exists]

    print(f"   - Itens para atualizar no Ultra: {len(to_update)}")
    print(f"   - Itens novos para adicionar: {len(to_add)}")

    if not to_update.empty:
        df_ultra.set_index("temp_key", inplace=True)
        to_update_indexed = to_update.set_index("temp_key")[cols_to_sync]
        df_ultra.update(to_update_indexed)
        df_ultra.reset_index(inplace=True)

    if not to_add.empty:
        df_ultra = pd.concat([df_ultra, to_add], ignore_index=True)

    df_ultra = ensure_valor_venda_total(df_ultra)

    if (
        "PARECER_JURIDICO_IMPUGNACAO" in df_ultra.columns
        and "PARECER_JURIDICO_IMPUGNACAO" not in original_cols
    ):
        idx = original_cols.index("JUSTIFICATIVA_TECNICA") + 1 if "JUSTIFICATIVA_TECNICA" in original_cols else len(original_cols)
        original_cols.insert(idx, "PARECER_JURIDICO_IMPUGNACAO")
        print("Adicionando 'PARECER_JURIDICO_IMPUGNACAO' apos 'JUSTIFICATIVA_TECNICA'.")

    if "VALOR_VENDA_TOTAL" in df_ultra.columns and "VALOR_VENDA_TOTAL" not in original_cols:
        original_cols = insert_col_after(original_cols, "VALOR_VENDA_TOTAL", "VALOR_VENDA")
        print("Adicionando 'VALOR_VENDA_TOTAL' apos 'VALOR_VENDA'.")

    df_ultra.drop(columns=["temp_key"], inplace=True, errors="ignore")

    final_cols = [col for col in original_cols if col in df_ultra.columns]
    remaining = [col for col in df_ultra.columns if col not in final_cols]
    df_ultra = df_ultra[final_cols + remaining]

    # --- SALVAMENTO PRESERVANDO ESTILOS (openpyxl) ---
    print(f"Salvando em {PATH_ULTRA} preservando formatacao...")
    wb = load_workbook(PATH_ULTRA)
    ws = wb.active

    # Limpar apenas os dados, mantendo estilos se possível (opcional, dependendo da necessidade)
    # Aqui vamos sobrescrever as células a partir da linha 2 (assumindo header na linha 1)
    
    # Escrever os dados do DataFrame a partir da célula A2
    # Nota: index=False do pandas equivale a não incluir o índice aqui
    rows = dataframe_to_rows(df_ultra, index=False, header=False)
    
    for r_idx, row in enumerate(rows, 2): # Começa na linha 2 (abaixo do header)
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Opcional: Remover linhas que sobraram se o novo DF for menor (raro neste script)
    if ws.max_row > len(df_ultra) + 1:
        ws.delete_rows(len(df_ultra) + 2, ws.max_row - (len(df_ultra) + 1))

    wb.save(PATH_ULTRA)
    print(f"Sincronização concluída. {len(df_ultra)} itens totais no Ultra.")
    return True


def executar_arquivamento():
    print("\n--- Iniciando arquivamento de editais ---")
    if not os.path.exists(SCRIPT_ARCHIVE):
        print("Aviso: script de arquivamento nao encontrado.")
        return

    try:
        result = subprocess.run(
            [sys.executable, SCRIPT_ARCHIVE],
            check=True,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="ignore",
        )
        print(result.stdout)
        print("Arquivamento finalizado.")
    except Exception as exc:
        print(f"Erro no arquivamento: {exc}")


def main():
    if concatenar_planilhas():
        # Descomente para rodar o arquivador automatico do Trello.
        # executar_arquivamento()
        pass


if __name__ == "__main__":
    main()
